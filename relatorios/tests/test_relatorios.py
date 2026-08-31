import io
from datetime import date, timedelta
from decimal import Decimal

import pytest
from django.core import mail
from django.utils import timezone

from almoxarifado import services as almox
from almoxarifado.models import Item
from apontamentos import services as ap_services
from apontamentos.models import MotivoRetrabalho, TipoTrabalho
from chamados import services as cham
from core.models import CentroCusto
from relatorios import export
from relatorios import selectors as rel

pytestmark = pytest.mark.django_db


def hoje_as(h):
    return timezone.localtime(timezone.now()).replace(hour=h, minute=0, second=0, microsecond=0)


@pytest.fixture
def cenario(usuarios, setores, categorias):
    dev = TipoTrabalho.objects.create(nome="Desenvolvimento", slug="desenvolvimento")
    esp = TipoTrabalho.objects.create(nome="Espera de terceiro", slug="espera_terceiro", contabiliza_capacidade=False)
    ret = TipoTrabalho.objects.create(nome="Retrabalho", slug="retrabalho", exige_causa=True)
    motivo = MotivoRetrabalho.objects.create(nome="Requisito incompleto")
    u = usuarios["colab_ti"]
    c = cham.abrir_chamado(solicitante=usuarios["colab_prd"], titulo="A", descricao="d", categoria=categorias["suporte"], prioridade="media")
    ap_services.criar_apontamento(usuario=u, tipo=dev, chamado=c, inicio=hoje_as(8), fim=hoje_as(11))
    ap_services.criar_apontamento(usuario=u, tipo=esp, chamado=c, inicio=hoje_as(11), fim=hoje_as(12))
    ap_services.criar_apontamento(usuario=u, tipo=ret, chamado=c, inicio=hoje_as(13), fim=hoje_as(14),
                                  motivo_retrabalho=motivo, detalhe_retrabalho="Requisito incompleto do solicitante")  # fmt: skip
    for p in ("triagem", "fila", "execucao", "testes", "entregue"):
        c = cham.transicionar(chamado=c, para=p, usuario=usuarios["ger_ti"])
    item = Item.objects.create(codigo="X-1", descricao="Item", unidade="UN", setor_dono=setores["MAN"], estoque_minimo=5, custo_unitario=Decimal("2.00"))
    almox.registrar_movimento(item=item, setor=setores["MAN"], tipo="entrada", quantidade=3, usuario=usuarios["admin"])  # 3 <= 5
    cc = CentroCusto.objects.create(codigo="9", descricao="x", setor=setores["MAN"])
    almox.registrar_movimento(item=item, setor=setores["MAN"], tipo="saida", quantidade=1, usuario=usuarios["admin"], centro_custo=cc)
    return c


def test_painel_confere_com_consulta_bruta(api, usuarios, cenario):
    r = api(usuarios["ger_ti"]).get("/api/v1/relatorios/painel/")
    assert r.status_code == 200
    k = r.data["kpis"]
    assert k["horas_apontadas"] == 5.0            # 3 + 1 + 1
    assert k["retrabalho_percentual"] == 20.0     # 1 / 5
    assert k["sla_percentual"] == 100.0 and k["sla_meta"] == 90
    assert k["itens_abaixo_minimo"] == 1 and k["setores_com_item_abaixo"] == 1
    assert r.data["mini"]["espera_terceiro_horas"] == 1.0
    assert r.data["mini"]["cobertura_documentacao"] == 0.0  # suporte entregue sem doc
    assert r.data["retrabalho_por_motivo"][0]["motivo"] == "Requisito incompleto"
    assert r.data["consumo_por_setor"][0]["setor"] == "MAN"
    assert {linha["tipo"] for linha in r.data["horas_por_tipo"]} == {"Desenvolvimento", "Espera de terceiro", "Retrabalho"}
    # Colaborador não tem painel na matriz
    assert api(usuarios["colab_prd"]).get("/api/v1/relatorios/painel/").status_code == 403


def test_horas_previstas_usa_capacidade_e_dias_uteis(usuarios):
    de, ate = date(2026, 8, 24), date(2026, 8, 28)  # seg–sex
    assert rel.dias_uteis(de, ate) == 5
    # fixture: colab_ti, ger_ti, admin no setor TI, 480 min cada
    assert rel.horas_previstas(de, ate) == 3 * 8 * 5


def test_relatorio_sla(api, usuarios, cenario):
    r = api(usuarios["ger_ti"]).get("/api/v1/relatorios/sla/")
    assert r.data["total"] == 1 and r.data["percentual"] == 100.0
    assert r.data["por_categoria"][0] == {"categoria": "Suporte", "total": 1, "cumpridos": 1, "percentual": 100.0}
    assert r.data["por_prioridade"][0]["prioridade"] == "media"


def test_relatorio_auditoria(api, usuarios, cenario):
    r = api(usuarios["ger_ti"]).get(f"/api/v1/relatorios/auditoria/?objeto=chamados.chamado:{cenario.id}&acao=chamado.")
    assert r.status_code == 200
    acoes = [x["acao"] for x in r.data["registros"]]
    assert "chamado.entregue" in acoes and "chamado.abrir" in acoes
    assert api(usuarios["colab_prd"]).get("/api/v1/relatorios/auditoria/").status_code == 403


def test_export_csv_e_xlsx(api, usuarios, cenario):
    r = api(usuarios["ger_ti"]).get("/api/v1/relatorios/horas/?formato=csv")
    assert r.status_code == 200 and r["Content-Type"].startswith("text/csv")
    assert "nome;sobrenome;setor;minutos" in r.content.decode("utf-8-sig").splitlines()[0]
    r = api(usuarios["ger_ti"]).get("/api/v1/relatorios/sla/?formato=xlsx")
    assert r.status_code == 200 and r["Content-Disposition"] == 'attachment; filename="sla.xlsx"'
    from openpyxl import load_workbook

    ws = load_workbook(io.BytesIO(r.content)).active
    assert [c.value for c in ws[1]] == ["categoria", "total", "cumpridos", "percentual"]
    assert ws["A2"].value == "Suporte"
    assert api(usuarios["ger_ti"]).get("/api/v1/relatorios/sla/?formato=doc").status_code == 400


def test_export_grande_vira_job_e_email(api, usuarios, cenario, monkeypatch):
    # `cenario` é o que gera auditoria: sem ele o relatório sai com zero linhas e
    # `len(linhas) > LIMITE_SINCRONO` nunca é verdade, mesmo com o limite em 0.
    monkeypatch.setattr(export, "LIMITE_SINCRONO", 0)
    r = api(usuarios["ger_ti"]).get("/api/v1/relatorios/auditoria/?formato=csv")
    assert r.status_code == 202 and "job_id" in r.data
    # CELERY_TASK_ALWAYS_EAGER: o e-mail já saiu
    assert len(mail.outbox) == 1 and mail.outbox[0].attachments[0][0] == "auditoria.csv"


def test_aprovacao_em_lote(api, usuarios, categorias):
    tipo = TipoTrabalho.objects.create(nome="Análise", slug="analise")
    c = cham.abrir_chamado(solicitante=usuarios["colab_prd"], titulo="A", descricao="d", categoria=categorias["suporte"], prioridade="media")
    ini = hoje_as(8) - timedelta(days=10)
    aps = [ap_services.criar_apontamento(usuario=usuarios["colab_prd"], tipo=tipo, chamado=c,
                                         inicio=ini + timedelta(hours=i * 2), fim=ini + timedelta(hours=i * 2 + 1)) for i in range(3)]  # fmt: skip
    assert all(a.pendente_aprovacao for a in aps)
    ger = api(usuarios["ger_prd"])
    r = ger.post("/api/v1/apontamentos/decidir-lote/", {"ids": [aps[0].id, aps[1].id], "aprovar": True}, format="json")
    assert r.status_code == 200 and r.data["decididos"] == [aps[0].id, aps[1].id]
    r = ger.post("/api/v1/apontamentos/decidir-lote/", {"ids": [aps[2].id], "aprovar": False, "motivo": "Sem evidência"}, format="json")
    assert r.status_code == 200
    aps[2].refresh_from_db()
    assert aps[2].recusado_em and aps[2].motivo_recusa == "Sem evidência" and not aps[2].pendente_aprovacao
    # recusado não entra nos indicadores; aprovados entram
    r = api(usuarios["ger_ti"]).get("/api/v1/relatorios/horas/")
    assert r.data["total_min"] == 120
    # lote com um já decidido → 409 e nada muda
    r = ger.post("/api/v1/apontamentos/decidir-lote/", {"ids": [aps[0].id], "aprovar": True}, format="json")
    assert r.status_code == 409
    # fora do escopo → 404
    r = ger.post("/api/v1/apontamentos/decidir-lote/", {"ids": [999999], "aprovar": True}, format="json")
    assert r.status_code == 404
